import requests
import json
from bs4 import BeautifulSoup
import bs4
from openpyxl import Workbook


class MoodleExtractor:

    def __init__(self, sessKey: str, MoodleSession: str, SiteUrl: str) -> None:
        """
        You can get the ``sessKey`` and ``MoodleSession`` data by running the "M.cfg.sesskey" and "document.cookie"
        command in your browser console.
        """
        self.sessKey = sessKey
        self.headers = {
            "Cookie": f"MoodleSession={MoodleSession}",
            "Content-Type": "application/json",
        }
        self.SiteUrl = SiteUrl

    def ExtractParticipants(self, courseCode: int) -> None:
        self.ParticipantsList = []
        self.courseCode = courseCode
        url = f"{self.SiteUrl}/lib/ajax/service.php?sesskey={self.sessKey}&info=core_table_get_dynamic_table_content"
        payload = json.dumps(
            [
                {
                    "index": 0,
                    "methodname": "core_table_get_dynamic_table_content",
                    "args": {
                        "component": "core_user",
                        "handler": "participants",
                        "uniqueid": "user-index-participants",
                        "sortdata": [{"sortby": "lastname", "sortorder": 4}],
                        "jointype": 0,
                        "filters": {
                            "courseid": {
                                "name": "courseid",
                                "jointype": 0,
                                "values": [courseCode],
                            }
                        },
                        "firstinitial": "",
                        "lastinitial": "",
                        "pagenumber": "1",
                        "pagesize": "5000",
                        "hiddencolumns": [],
                        "resetpreferences": True,
                    },
                }
            ]
        )
        response = requests.request(
            "POST", url, headers=self.headers, data=payload, verify=False
        )
        js = json.loads(response.text)
        soup = BeautifulSoup(js[0]["data"]["html"], "html.parser")
        participants_table = soup.find("table", id="participants")
        participants_tbody = participants_table.find("tbody")
        participants_tr = participants_tbody.find_all("tr")
        for participant in participants_tr:
            if "emptyrow" in participant.get("class"):
                break
            participant: bs4.element.Tag
            participant_th = participant.find("th")
            name_surname = participant_th.find("a").text
            self.ParticipantsList.append(name_surname)
        print("Participants Extracted Successfully")

    def SaveToTable(self):
        """
        Save participants to table file.
        """
        wb = Workbook()
        ws = wb.active
        count = 2
        for participant in self.ParticipantsList:
            ws["A1"] = "NAME SURNAME"
            ws[f"A{str(count)}"] = participant
            count += 1
        file_name = f"{str(self.courseCode)}.xlsx"
        wb.save(file_name)

    def SaveToJson(self):
        """
        Save participants to JSON file.
        """
        data = {"participants": self.ParticipantsList}
        with open(f"{self.courseCode}.json", "w", encoding="utf-8") as f:
            f.write(json.dumps(data, ensure_ascii=False))
